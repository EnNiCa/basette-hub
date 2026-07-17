import os
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, current_app, abort, Response, send_file
from extensions import get_db
from auth.decorators import login_requerido, solo_admin
from auth.permisos import ids_visibles
from utils.validaciones import dni_valido, telefono_valido, cups_valido, iban_valido, cp_valido, cif_valido
from werkzeug.utils import secure_filename
from datetime import date

ventas_bp = Blueprint('ventas', __name__, url_prefix='/ventas')

MODULOS_VALIDOS = ('energia', 'alarmas', 'telefonia', 'placas_solares')
ESTADOS_VALIDOS = ('nulo', 'pendiente_carga', 'pendiente_firma', 'scoring',
                   'activacion', 'activa', 'incidencia', 'baja')

CATEGORIAS_ARCHIVO = ('dni', 'certificado_bancario', 'escritura', 'justo_titulo',
                      'cif', 'facturas', 'acta_comunidad', 'otros')


@ventas_bp.route('/general')
@login_requerido
def listar_ventas_general():
    buscar = request.args.get('buscar', '').strip()
    ventas = buscar_ventas(None, buscar)
    es_admin = session['rol'] == 'admin'

    return render_template(
        'ventas/listado_general.html',
        ventas=ventas,
        es_admin=es_admin,
        modulo_actual=None,
        buscar=buscar
    )

@ventas_bp.route('/')
@login_requerido
def listar_ventas():
    modulo_filtro = request.args.get('modulo', 'energia')
    if modulo_filtro not in MODULOS_VALIDOS:
        modulo_filtro = 'energia'

    buscar = request.args.get('buscar', '').strip()
    ventas = buscar_ventas(modulo_filtro, buscar)
    es_admin = session['rol'] == 'admin'

    return render_template(
        f'ventas/listado_{modulo_filtro}.html',
        ventas=ventas,
        es_admin=es_admin,
        modulo_actual=modulo_filtro,
        buscar=buscar
    )
    
@ventas_bp.route('/<int:venta_id>/estado', methods=['POST'])
@solo_admin
def actualizar_estado(venta_id):
    nuevo_estado = request.form.get('estado')

    if nuevo_estado not in ESTADOS_VALIDOS:
        flash('Estado no válido.', 'error')
        return redirect(url_for('ventas.listar_ventas'))

    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        "UPDATE ventas SET estado = %s WHERE id = %s",
        (nuevo_estado, venta_id)
    )
    db.commit()

    flash('Estado actualizado correctamente.', 'exito')

    modulo_actual = request.form.get('modulo_actual', 'energia')
    return redirect(url_for('ventas.listar_ventas', modulo=modulo_actual))

@ventas_bp.route('/nueva', methods=['GET', 'POST'])
@login_requerido
def nueva_venta():
    db = get_db()
    cursor = db.cursor(dictionary=True)

    if request.method == 'POST':
        dni = request.form.get('dni', '').strip().upper()
        cif = request.form.get('cif', '').strip().upper()
        tipo_cliente = request.form.get('tipo_cliente', 'particular')
        razon_social = request.form.get('razon_social', '').strip()
        telefono = request.form.get('telefono', '').strip()
        cups = request.form.get('cups', '').strip().upper()
        numero_cuenta = request.form.get('numero_cuenta', '').strip().upper()
        cp = request.form.get('cp', '').strip()
        modulo = request.form.get('modulo')

        errores = []
        if tipo_cliente == 'particular' and not dni_valido(dni):
            errores.append('El DNI/NIE debe tener un formato válido (ej: 12345678A o X1234567A).')
        if tipo_cliente == 'empresa' and not cif_valido(cif):
            errores.append('El CIF debe tener un formato válido (ej: B12345678).')
        if tipo_cliente == 'empresa' and not razon_social:
            errores.append('Debes indicar la razón social de la empresa.')
        if tipo_cliente == 'empresa' and not dni_valido(dni):
            errores.append('El DNI del representante debe tener un formato válido.')
        if telefono and not telefono_valido(telefono):
            errores.append('El teléfono debe tener exactamente 9 dígitos.')
        if not cups_valido(cups):
            errores.append('El CUPS no tiene un formato válido (ej: ES0021000005311232MT).')
        if not iban_valido(numero_cuenta):
            errores.append('El número de cuenta no tiene un formato IBAN válido.')
        if not cp_valido(cp):
            errores.append('El código postal debe tener 5 dígitos.')

        if errores:
            for error in errores:
                flash(error, 'error')
            cursor.execute("SELECT id, nombre, tipo_servicio FROM companias")
            companias = cursor.fetchall()
            cursor.execute("SELECT id, nombre, tipo_servicio FROM tarifas WHERE vigente = TRUE")
            tarifas = cursor.fetchall()
            cursor.execute("SELECT id, nombre FROM canales")
            canales = cursor.fetchall()
            return render_template(
                'ventas/nueva.html',
                companias=companias,
                tarifas=tarifas,
                canales=canales,
                valores=request.form
            )

        cursor.execute(
            """
            INSERT INTO ventas
                (comercial_id, modulo, tipo_energia, compania_id, tarifa_id, canal_id,
                nombre, apellidos, tipo_cliente, direccion, cp, dni, cif, razon_social,
                cups, telefono, email, numero_cuenta, observaciones, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                session['usuario_id'],
                modulo,
                request.form.get('tipo_energia') or None,
                request.form.get('compania_id'),
                request.form.get('tarifa_id'),
                request.form.get('canal_id') or None,
                request.form.get('nombre'),
                request.form.get('apellidos'),
                tipo_cliente,
                request.form.get('direccion'),
                cp or None,
                dni,
                cif or None,
                razon_social or None,
                cups or None,
                telefono or None,
                request.form.get('email'),
                numero_cuenta or None,
                request.form.get('observaciones'),
                'pendiente_carga',
            )
        )
        db.commit()
        return redirect(url_for('ventas.listar_ventas', modulo=modulo))

    cursor.execute("SELECT id, nombre, tipo_servicio FROM companias")
    companias = cursor.fetchall()

    cursor.execute("SELECT id, nombre, tipo_servicio FROM tarifas WHERE vigente = TRUE")
    tarifas = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM canales")
    canales = cursor.fetchall()

    return render_template(
        'ventas/nueva.html',
        companias=companias,
        tarifas=tarifas,
        canales=canales
    )
    
@ventas_bp.route('/<int:venta_id>/archivos')
@login_requerido
def ver_archivos(venta_id):
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT id, comercial_id, modulo, nombre, apellidos, razon_social FROM ventas WHERE id = %s", (venta_id,))
    venta = cursor.fetchone()
    if venta is None:
        abort(404)

    visibles = ids_visibles(session['usuario_id'], session['rol'])
    if visibles is not None and venta['comercial_id'] not in visibles:
        abort(403)

    cursor.execute(
        "SELECT id, categoria, nombre_archivo, fecha_subida FROM venta_archivos WHERE venta_id = %s",
        (venta_id,)
    )
    archivos_todos = cursor.fetchall()

    archivos_por_categoria = {cat: [] for cat in CATEGORIAS_ARCHIVO}
    for archivo in archivos_todos:
        archivos_por_categoria[archivo['categoria']].append(archivo)

    origen = request.args.get('origen') or url_for('ventas.listar_ventas', modulo=venta['modulo'])

    return render_template(
        'ventas/archivos_adjuntos.html',
        venta=venta,
        archivos_por_categoria=archivos_por_categoria,
        origen=origen
    )


@ventas_bp.route('/<int:venta_id>/archivo', methods=['POST'])
@login_requerido
def subir_archivo(venta_id):
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT comercial_id FROM ventas WHERE id = %s", (venta_id,))
    venta = cursor.fetchone()
    if venta is None:
        abort(404)

    visibles = ids_visibles(session['usuario_id'], session['rol'])
    if visibles is not None and venta['comercial_id'] not in visibles:
        abort(403)

    categoria = request.form.get('categoria')
    if categoria not in CATEGORIAS_ARCHIVO:
        flash('Categoría de archivo no válida.', 'error')
        return redirect(url_for('ventas.ver_archivos', venta_id=venta_id))

    archivos = request.files.getlist('archivo')

    if not archivos or archivos[0].filename == '':
        flash('No se ha seleccionado ningún archivo.', 'error')
        return redirect(url_for('ventas.ver_archivos', venta_id=venta_id))

    for archivo in archivos:
        if archivo.filename == '':
            continue
        nombre_seguro = secure_filename(archivo.filename)
        nombre_final = f"{venta_id}_{categoria}_{nombre_seguro}"
        ruta_completa = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre_final)
        archivo.save(ruta_completa)

        cursor.execute(
            "INSERT INTO venta_archivos (venta_id, categoria, nombre_archivo, ruta_archivo) VALUES (%s, %s, %s, %s)",
            (venta_id, categoria, archivo.filename, ruta_completa)
        )

    db.commit()
    flash('Archivo(s) subido(s) correctamente.', 'exito')
    return redirect(url_for('ventas.ver_archivos', venta_id=venta_id))

@ventas_bp.route('/<int:venta_id>/editar', methods=['GET', 'POST'])
@login_requerido
def editar_venta(venta_id):
    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute("SELECT * FROM ventas WHERE id = %s", (venta_id,))
    venta = cursor.fetchone()
    if venta is None:
        abort(404)

    visibles = ids_visibles(session['usuario_id'], session['rol'])
    if visibles is not None and venta['comercial_id'] not in visibles:
        abort(403)

    origen = request.args.get('origen') or request.form.get('origen') or url_for('ventas.listar_ventas', modulo=venta['modulo'])

    if request.method == 'POST':
        dni = request.form.get('dni', '').strip().upper()
        cif = request.form.get('cif', '').strip().upper()
        tipo_cliente = request.form.get('tipo_cliente', 'particular')
        razon_social = request.form.get('razon_social', '').strip()
        telefono = request.form.get('telefono', '').strip()
        cups = request.form.get('cups', '').strip().upper()
        numero_cuenta = request.form.get('numero_cuenta', '').strip().upper()
        cp = request.form.get('cp', '').strip()

        errores = []
        if tipo_cliente == 'particular' and not dni_valido(dni):
            errores.append('El DNI/NIE debe tener un formato válido (ej: 12345678A o X1234567A).')
        if tipo_cliente == 'empresa' and not cif_valido(cif):
            errores.append('El CIF debe tener un formato válido (ej: B12345678).')
        if tipo_cliente == 'empresa' and not razon_social:
            errores.append('Debes indicar la razón social de la empresa.')
        if tipo_cliente == 'empresa' and not dni_valido(dni):
            errores.append('El DNI del representante debe tener un formato válido.')
        if telefono and not telefono_valido(telefono):
            errores.append('El teléfono debe tener exactamente 9 dígitos.')
        if not cups_valido(cups):
            errores.append('El CUPS no tiene un formato válido (ej: ES0021000005311232MT).')
        if not iban_valido(numero_cuenta):
            errores.append('El número de cuenta no tiene un formato IBAN válido.')
        if not cp_valido(cp):
            errores.append('El código postal debe tener 5 dígitos.')

        if errores:
            for error in errores:
                flash(error, 'error')
            cursor.execute(
                "SELECT id, nombre FROM companias WHERE tipo_servicio = %s", (venta['modulo'],)
            )
            companias = cursor.fetchall()
            cursor.execute(
                "SELECT id, nombre FROM tarifas WHERE tipo_servicio = %s AND vigente = TRUE", (venta['modulo'],)
            )
            tarifas = cursor.fetchall()
            cursor.execute("SELECT id, nombre FROM canales")
            canales = cursor.fetchall()
            venta.update(request.form.to_dict())
            venta['id'] = venta_id
            return render_template('ventas/editar.html', venta=venta, companias=companias, tarifas=tarifas, canales=canales, origen=origen)

        cursor.execute(
            """
            UPDATE ventas SET
                tipo_energia = %s, compania_id = %s, tarifa_id = %s, canal_id = %s,
                nombre = %s, apellidos = %s, tipo_cliente = %s, direccion = %s, cp = %s,
                dni = %s, cif = %s, razon_social = %s, cups = %s,
                telefono = %s, email = %s, numero_cuenta = %s, observaciones = %s
            WHERE id = %s
            """,
            (
                request.form.get('tipo_energia') or None,
                request.form.get('compania_id'),
                request.form.get('tarifa_id'),
                request.form.get('canal_id') or None,
                request.form.get('nombre'),
                request.form.get('apellidos'),
                tipo_cliente,
                request.form.get('direccion'),
                cp or None,
                dni,
                cif or None,
                razon_social or None,
                cups or None,
                telefono or None,
                request.form.get('email'),
                numero_cuenta or None,
                request.form.get('observaciones'),
                venta_id,
            )
        )
        db.commit()
        flash('Venta actualizada correctamente.', 'exito')
        return redirect(origen)

    cursor.execute(
        "SELECT id, nombre FROM companias WHERE tipo_servicio = %s", (venta['modulo'],)
    )
    companias = cursor.fetchall()
    cursor.execute(
        "SELECT id, nombre FROM tarifas WHERE tipo_servicio = %s AND vigente = TRUE", (venta['modulo'],)
    )
    tarifas = cursor.fetchall()
    cursor.execute("SELECT id, nombre FROM canales")
    canales = cursor.fetchall()

    return render_template('ventas/editar.html', venta=venta, companias=companias, tarifas=tarifas, canales=canales, origen=origen)

def buscar_ventas(modulo_filtro=None, buscar=''):
    db = get_db()
    cursor = db.cursor(dictionary=True)
    visibles = ids_visibles(session['usuario_id'], session['rol'])
    
    sql = """
        SELECT v.*, c.nombre AS compania, t.nombre AS tarifa, u.nombre AS comercial,
               ca.nombre AS canal, b.fecha_baja AS fecha_baja
        FROM ventas v
        JOIN companias c ON v.compania_id = c.id
        JOIN tarifas t ON v.tarifa_id = t.id
        JOIN usuarios u ON v.comercial_id = u.id
        LEFT JOIN canales ca ON v.canal_id = ca.id
        LEFT JOIN bajas b ON v.id = b.venta_id
    """
    
    condiciones = []
    parametros = []
    
    if modulo_filtro:
        condiciones.append("v.modulo = %s")
        parametros.append(modulo_filtro)
        
    if visibles is not None:
        placeholders = ','.join(['%s'] * len(visibles))
        condiciones.append(f"v.comercial_id IN ({placeholders})")
        parametros += visibles
        
    if buscar:
        condiciones.append("""(
            CONCAT(v.nombre, ' ', v.apellidos) LIKE %s
            OR v.dni LIKE %s
            OR v.telefono LIKE %s
            OR v.email LIKE %s
            OR v.cups LIKE %s
        )""")
        comodin = f"%{buscar}%"
        parametros += [comodin, comodin, comodin, comodin, comodin]
        
    if condiciones:
        sql += " WHERE " + " AND ".join(condiciones)
        
    sql += " ORDER BY v.id DESC"
    
    cursor.execute(sql, tuple(parametros))
    return cursor.fetchall()

def columnas_export(es_admin):
    columnas = [
        ('nombre', 'Nombre', 'texto'),
        ('apellidos', 'Apellidos', 'texto'),
        ('dni', 'DNI', 'texto'),
        ('direccion', 'Dirección', 'texto'),
        ('cp', 'C.P.', 'texto'),
        ('telefono', 'Teléfono', 'texto'),
        ('email', 'Email', 'texto'),
        ('numero_cuenta', 'Nº Cuenta', 'texto'),
        ('comercial', 'Comercial', 'texto'),
        ('canal', 'Canal', 'texto'),
        ('modulo', 'Módulo', 'texto'),
        ('tipo_energia', 'Tipo energía', 'texto'),
        ('compania', 'Compañía', 'texto'),
        ('tarifa', 'Tarifa', 'texto'),
        ('cups', 'CUPS', 'texto'),
        ('mantenimiento', 'Mantenimiento', 'bool'),
        ('bateria', 'Batería', 'bool'),
        ('estado', 'Estado', 'texto'),
        ('fecha_firma', 'Fecha firma', 'fecha'),
        ('fecha_activacion', 'Fecha activación', 'fecha'),
        ('fecha_baja', 'Fecha baja', 'fecha'),
    ]
    if es_admin:
        columnas += [
            ('fecha_liquidacion', 'Fecha liquidación', 'fecha'),
            ('importe_liquidar', 'Importe compañía', 'texto'),
            ('fecha_descomision', 'Fecha descomisión', 'fecha'),
            ('importe_descomisionado', 'Importe descomisión', 'texto'),
        ]
    columnas += [
        ('fecha_pago_comercial', 'Fecha pago comercial', 'fecha'),
        ('importe_pago_comercial', 'Importe pago comercial', 'texto'),
        ('fecha_descomision_comercial', 'Fecha descomisión comercial', 'fecha'),
        ('importe_descomisionado_comercial', 'Importe descomisión comercial', 'texto'),
        ('observaciones', 'Observaciones', 'texto'),
    ]
    return columnas

def formatear_valor(valor, tipo):
    if valor is None:
        return ''
    if tipo == 'bool':
        return 'Sí' if valor else 'No'
    if tipo == 'fecha' and isinstance(valor, date):
        return valor.strftime('%d/%m/%Y')
    return valor

@ventas_bp.route('/exportar/csv')
@login_requerido
def exportar_csv():
    modulo_filtro = request.args.get('modulo') or None
    if modulo_filtro and modulo_filtro not in MODULOS_VALIDOS:
        modulo_filtro = None
    buscar = request.args.get('buscar', '').strip()

    ventas = buscar_ventas(modulo_filtro, buscar)
    es_admin = session['rol'] == 'admin'
    columnas = columnas_export(es_admin)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow([etiqueta for clave, etiqueta, tipo in columnas])
    for venta in ventas:
        writer.writerow([formatear_valor(venta.get(clave), tipo) for clave, etiqueta, tipo in columnas])

    respuesta = Response('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8')
    nombre_archivo = f"ventas_{modulo_filtro or 'general'}.csv"
    respuesta.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
    return respuesta


@ventas_bp.route('/exportar/excel')
@login_requerido
def exportar_excel():
    modulo_filtro = request.args.get('modulo') or None
    if modulo_filtro and modulo_filtro not in MODULOS_VALIDOS:
        modulo_filtro = None
    buscar = request.args.get('buscar', '').strip()

    ventas = buscar_ventas(modulo_filtro, buscar)
    es_admin = session['rol'] == 'admin'
    columnas = columnas_export(es_admin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append([etiqueta for clave, etiqueta, tipo in columnas])
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    for venta in ventas:
        ws.append([formatear_valor(venta.get(clave), tipo) for clave, etiqueta, tipo in columnas])

    for columna in ws.columns:
        valores = [len(str(celda.value)) for celda in columna if celda.value]
        ancho = max(valores) + 2 if valores else 12
        ws.column_dimensions[columna[0].column_letter].width = min(ancho, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"ventas_{modulo_filtro or 'general'}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )